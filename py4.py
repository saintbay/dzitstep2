from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import time
from openpyxl import Workbook

chrome = webdriver.Chrome(keep_alive=True)

chrome.get('https://www.saucedemo.com/')

username = chrome.find_element(By.ID, 'user-name').send_keys('standard_user')
password = chrome.find_element(By.NAME, 'password').send_keys('secret_sauce')
button = chrome.find_element(By.XPATH, '//*[@id="login-button"]').click()
img = chrome.find_elements(By.CLASS_NAME, 'inventory_item_img')
description = chrome.find_elements(By.CLASS_NAME, 'inventory_item_desc')
name = chrome.find_elements(By.CLASS_NAME, 'inventory_item_name')
price = chrome.find_elements(By.CLASS_NAME, 'inventory_item_price')

chrome.page_source

wb = Workbook()
ws = wb.active
ws['A1'] = 'Numeration'
ws['B1'] = 'Image'
ws['C1'] = 'Description'
ws['D1'] = 'Name'
ws['E1'] = 'Price'

for i in range(1, 7):
    ws[f'A{i+1}'] = i
    ws[f'C{i+1}'] = description[i-1].text
    ws[f'D{i+1}'] = name[i-1].text
    ws[f'E{i+1}'] = price[i-1].text

row = 2
for i in range(1,12,2):
    ws[f'B{row}'] = img[i].get_attribute('src')
    row += 1

wb.save("example.xlsx")
time.sleep(10)